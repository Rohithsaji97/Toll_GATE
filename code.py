import cv2 as cv
import pytesseract
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#imported required libraries

pytesseract.pytesseract.tesseract_cmd = r'C:/Users/ASUS/AppData/Local/Programs/Tesseract-OCR/tesseract.exe'
#assigning tesseract path


def img_capt():
    #initialize the camera,here we are using laptop web cam
    cam=cv.VideoCapture(0)    
    result,image=cam.read()  
    #cam.read() function gives 2 outputs boolean statement and numpy array
    del cam
    #turns off the camera        
    if result:
        cv.imshow("Present_number_plate", image)
        #shows the captured image in another window
        cv.imwrite("Present_number_plate.jpg", image)
        #saves the captured image in the given location
        cv.waitKey(3000)
        #shows the window for 3sec
        cv.destroyWindow("Present_number_plate")
        #the window is removed
    else:
        print('image not detected')
    return image


def img_txt(img):
    gray=cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    #converting to gray scale
    thresh=cv.threshold(gray, 0, 255, cv.THRESH_BINARY_INV + cv.THRESH_OTSU)[1]
    # Apply thresholding to convert to binary image
    kernel = cv.getStructuringElement(cv.MORPH_RECT, (3,3))
    cleaned = cv.morphologyEx(thresh, cv.MORPH_CLOSE, kernel)
    # Apply morphological operations to remove noise and enhance text
    text = pytesseract.image_to_string(cleaned, config='--psm 11')
    # Extract text using Tesseract OCR
    m=len(text)
    text=text[0:m-1]
    return text

        
def toll_amt(vn):
    #the directory is fixed
    os.chdir('C:/Users/ASUS/OneDrive/Desktop/Projects/Fasttag system')
    #flag is assigned
    Flag=False
    #workbook is loaded(an external workbook was created earlier)
    wb=load_workbook('vehicle_number.xlsx')
    #worksheet is loaded
    ws=wb.active
    char1=get_column_letter(1)
    char2=get_column_letter(2)
    #checking for the same number and 40 Rs is deducted
    for i in range (2,5):
        if vn==ws[char1+str(i)].value:
            ws[char2+str(i)]=(ws[char2+str(i)].value)-40
            bal=ws[char2+str(i)].value
            Flag=True
            wb.save('vehicle_number.xlsx')
            break
    if Flag:
        detail=bal
    else:
        detail='Vehicle has no sufficient balance'
    return detail
    

img=img_capt()
vehicle_number=img_txt(img)
print(vehicle_number)
bal_amt=toll_amt(vehicle_number)
print(bal_amt)
